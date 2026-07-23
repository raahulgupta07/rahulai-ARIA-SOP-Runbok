"""People / adoption Excel export — one .xlsx workbook the management dashboard
downloads. Pure read: builds entirely off productivity.overview(...) so the
numbers always match the on-screen scorecard. openpyxl is imported lazily so the
rest of the app never hard-depends on it."""
import datetime
import re
from io import BytesIO

from . import productivity

# per-user columns shared by every department sheet + the "All users" sheet
_COLUMNS = [
    ("Name", "name", 22),
    ("Email", "email", 30),
    ("Department", "dept", 18),
    ("Created", "created_at", 12),
    ("Last active", "last_active", 12),
    ("Active days", "active_days", 12),
    ("Questions", "questions", 11),
    ("Hours saved", "hours", 12),
    ("Value USD", "value_usd", 12),
    ("Tokens", "tokens", 12),
    ("Helpful %", "helpful", 11),
    ("Resolved %", "resolved", 11),
    ("Login", "_login", 16),
    ("Flag", "flag", 10),
]


def _sanitize_sheet(name: str) -> str:
    """Excel sheet titles: max 31 chars, none of []:*?/\\ — truncate to 28."""
    clean = re.sub(r"[\[\]:*?/\\]", " ", str(name or "")).strip() or "Sheet"
    return clean[:28]


def _cell_value(person: dict, key: str):
    """Map a people[] row onto one column's value, keeping numbers as numbers."""
    if key == "_login":
        return ", ".join(person.get("methods") or [])
    if key == "helpful":
        v = person.get("helpful")
        return v if v is not None else ""
    if key == "resolved":
        # productivity stores resolved as a 0..1 rate → show as a percent number
        v = person.get("resolved")
        return round(float(v) * 100, 1) if v is not None else ""
    v = person.get(key)
    return "" if v is None else v


def people_xlsx(days: int = 30, date_from: str | None = None,
                date_to: str | None = None) -> bytes:
    """Build the people workbook and return its bytes. Raises RuntimeError if
    openpyxl isn't available in this environment."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    data = productivity.overview(days=days, date_from=date_from, date_to=date_to)
    rng = data.get("range", {})
    totals = data.get("totals", {})
    depts = data.get("departments", [])
    domains = data.get("domains", [])
    people = data.get("people", [])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3F51B5")
    title_font = Font(bold=True, size=13)
    label_font = Font(bold=True)
    center = Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    def _style_header(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    # ---- Summary sheet --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "People & Adoption report"
    ws["A1"].font = title_font
    ws["A3"] = "Window"
    ws["A3"].font = label_font
    ws["B3"] = f"{rng.get('from', '?')} → {rng.get('to', '?')} ({rng.get('days', days)} days)"

    r = 5
    ws.cell(row=r, column=1, value="Totals").font = title_font
    r += 1
    total_rows = [
        ("Questions", totals.get("questions", 0)),
        ("Hours saved", totals.get("hours_saved", 0)),
        ("Labour value USD", totals.get("labour_value", 0)),
        ("AI cost USD", totals.get("ai_cost", 0)),
        ("ROI", totals.get("roi")),
        ("Resolved rate %", round(float(totals.get("resolved_rate", 0)) * 100, 1)),
    ]
    for label, val in total_rows:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value="" if val is None else val)
        r += 1

    # departments block
    r += 1
    ws.cell(row=r, column=1, value="By department").font = title_font
    r += 1
    dhead = ["Department", "Users", "Questions", "Hours saved"]
    for i, h in enumerate(dhead, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(dhead))
    r += 1
    for d in depts:
        ws.cell(row=r, column=1, value=d.get("dept"))
        ws.cell(row=r, column=2, value=d.get("users", 0))
        ws.cell(row=r, column=3, value=d.get("questions", 0))
        ws.cell(row=r, column=4, value=d.get("hours", 0))
        r += 1

    # domains block
    r += 1
    ws.cell(row=r, column=1, value="By domain").font = title_font
    r += 1
    mhead = ["Domain", "Users", "Questions", "Hours saved"]
    for i, h in enumerate(mhead, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(mhead))
    r += 1
    for d in domains:
        ws.cell(row=r, column=1, value=d.get("domain"))
        ws.cell(row=r, column=2, value=d.get("users", 0))
        ws.cell(row=r, column=3, value=d.get("questions", 0))
        ws.cell(row=r, column=4, value=d.get("hours", 0))
        r += 1

    for col, width in (("A", 22), ("B", 16), ("C", 14), ("D", 14)):
        ws.column_dimensions[col].width = width

    # ---- one sheet per department + the All-users sheet -----------------------
    def _write_people_sheet(title, rows):
        sh = wb.create_sheet(_sanitize_sheet(title))
        for i, (label, _key, width) in enumerate(_COLUMNS, start=1):
            sh.cell(row=1, column=i, value=label)
            sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        _style_header(sh, 1, len(_COLUMNS))
        rn = 2
        for person in rows:
            for i, (_label, key, _w) in enumerate(_COLUMNS, start=1):
                sh.cell(row=rn, column=i, value=_cell_value(person, key))
            rn += 1
        sh.freeze_panes = "A2"
        return sh

    # group people by department, preserving the questions-desc order within each
    by_dept: dict[str, list] = {}
    for p in people:
        by_dept.setdefault(p.get("dept") or "Unassigned", []).append(p)

    used = {"summary"}
    for dept_name in sorted(by_dept):
        base = _sanitize_sheet(dept_name)
        title = base
        n = 2
        while title.lower() in used:  # avoid duplicate sheet titles after truncation
            title = f"{base[:26]}~{n}"
            n += 1
        used.add(title.lower())
        _write_people_sheet(title, by_dept[dept_name])

    _write_people_sheet("All users", people)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
